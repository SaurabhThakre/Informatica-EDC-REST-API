'''
Script takes "tables.xlsx" file as input and generate "AssetsExport.csv" as output
Make sure above input file is present and Run below command to generate output file:
python fetchAssetDetails.py
'''
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
import html2text
import openpyxl

requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

# This function is for REST API call along with Authorization details and return JSON response
def getJSON(q_val):
    search_url= '<EDC_Catalog_URL>/access/2/catalog/data/objects'
    username='<AD_Group>/<Service_Account>'
    password='<Password>'

    payload={
        'q':q_val,
        'includeDstLinks':'false',
        'includeRefObjects':'false',
        'includeSrcLinks':'false',
        'pageSize':10000
    }
            
    headers = {
      'Content-Type': 'application/json',
      'Accept': 'application/json',
      # ***Important***: Generate below Auth key using encodeUser.py
      'Authorization': 'Basic <Add your Auth key here>'
    }
    
    return requests.get(search_url,auth=HTTPBasicAuth(username, password),params = payload, headers=headers,verify=False)
    
    
# This function fetches business description, Data owner, Data steward, Ad group, Data subject area  
# and Url details from JSON response and appends it to their respective lists
def getFacts(facts):
    nDescription,nOwner,nSteward,nAD,nSubject,nUrl = 0,0,0,0,0,0

    for i in facts:
        if "com.infa.ldm.ootb.enrichments.businessDescription" == i["attributeId"]:
            descriptions.append(html2text.html2text(i['value']))
            nDescription = 1
        elif "com.infa.ldm.ootb.enrichments.dataOwner" == i["attributeId"]:
            owners.append(i['value'])
            nOwner = 1
        elif "com.infa.ldm.ootb.enrichments.dataSteward" == i["attributeId"]:
            stewards.append(i['value'])
            nSteward = 1
        elif "AD Group" == i["label"]:
            adgroups.append(i['value'])
            nAD = 1
        elif "Data Subject Area" == i["label"]:
            subjects.append(i['value'])
            nSubject = 1
        elif "com.infa.ldm.ootb.enrichments.URL" == i["attributeId"]:
            URLs.append(i['value'])
            nUrl = 1
        elif "core.name" == i["attributeId"]:
            names.append(i['value'])

    if nDescription == 0:
        descriptions.append(None)
    if nOwner == 0:
        owners.append(None)
    if nSteward == 0:
        stewards.append(None)
    if nAD == 0:
        adgroups.append(None)
    if nSubject == 0:
        subjects.append(None)
    if nUrl == 0:
        URLs.append(None)
       
def checkTableName(tableObj):
    idArr = tableObj['id'].split('/')
    if idArr[-1] in inputTables and idArr[-1] not in tablePresent:
        tablePresent.append(idArr[-1])
        return True
    return False


# This function fetches Column JSON response and collects Table details in lists
def columnDetailsFromJson():
    q_vals = []

    # Format value for 'q' string, For example :- 
    # q=(id:(<Resource_name>\:<//Database_name/Schema_name/Table1>/* OR <Resource_name>\:<//Database_name/Schema_name/Table2>/* OR ...) AND core.classType:"com.infa.ldm.relational.Column")
    # DBRKS_SC_PRD_ADP_DSCRY_1://hive_metastore/prd_adp/thunderbird -> DBRKS_SC_PRD_ADP_DSCRY_1\://hive_metastore/prd_adp/thunderbird/*
    for i in inputIds:
        tid = i.split(':')
        val = 'id:{}\:{}/*'.format(tid[0],tid[1])
        q_vals.append(val)
      
    joined = " OR ".join(q_vals)
    q_val = '({}) AND core.classType:"com.infa.ldm.relational.Column"'.format(joined)
    
    # Calls function getJSON() and passes q string as parameter
    print("REST API call for collecting column details...")
    get_col_response = getJSON(q_val)
    getvalueJson = get_col_response.json()
    
    if get_col_response.status_code == 200:
        print("Response received successfully!")
    else:
        print("Response failed!")
        print(getvalueJson)
    
    colList = getvalueJson['items']
    
    print("Collecting column details...")
    # Iterate over all the columns
    for i in colList:
        # Add column id, type and schema name in their respective lists
        ids.append(i['id'])
        types.append('Column')
        
        # Add schema name for column in schemas lists
        idArr = i['id'].split('/')
        schemas.append(idArr[-3])
        
        facts = i['facts']
        # Calls getFacts() function and passes facts object as parameter
        # facts object contains all the details of Asset along with attributes data
        getFacts(facts)
    print("Collected column details successfully.") 

# This function fetches Table JSON response and collects Table details in lists
def tableDetailsFromJson():
    
    # Format value for 'q' string, For example :- 
    # q=(core.classType:"com.infa.ldm.relational.Table" AND core.name:("<Table1>" OR <Table2> OR ... OR "<Tablen>"))
    joined = " OR ".join(inputTablesFormatted)
    q_val = 'core.classType:"com.infa.ldm.relational.Table" AND core.name:({})'.format(joined)
    
    # Calls function getJSON() and passes q string as parameter
    print("Sent REST API call for collecting table details...")
    get_response = getJSON(q_val)
    getvalueJson = get_response.json()
    
    if get_response.status_code == 200:
        print("Response received successfully!")
    else:
        print("Response failed!")
        print(getvalueJson)
    
    # Response received from JSON object is internally sorted on basis of core.name, core.lastModified
    # Sort Tables Json object based on number of information available in descending order
    getvalueJson['items'].sort(key = lambda x: len(x['facts']), reverse=True)

    print("Collecting table details...")
    for i in getvalueJson['items']:
        if checkTableName(i):
            # Add table id to inputIds lists to later user for column q_val
            inputIds.append(i['id'])
            
            # Add table id and type in their respective lists
            ids.append(i['id'])
            types.append('Table')
    
            # Add schema name for table in schemas lists
            idArr = i['id'].split('/')
            schemas.append(idArr[-2])
    
            facts = i['facts']
            # Calls getFacts() function and passes facts object as parameter
            # facts object contains all the details of Asset along with attributes data
            getFacts(facts)
    print("Collected table details successfully.") 
    
    # Calls columnDetailsFromJson() function to fetch columns detail of all tables
    columnDetailsFromJson()

# Start ------------------------------------------------------------------------------------------------------

# Define lists to collect Table and Column details
data = {}
inputTables = []
inputTablesFormatted = []
inputIds = []
tablePresent = []
ids = []
names = []
types = []
schemas = []
descriptions = []
owners = []
stewards = []
adgroups = []
subjects = []
URLs = []

print("Reading input file...")

# Define variable to load the dataframe
# Take excel file as input for list of tables
dataframe = openpyxl.load_workbook("tablesNonProd.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values (i.e., Table names)
for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        if(not None):
            # Add table in inputTables[] list
            inputTables.append(col[row].value)
            inputTablesFormatted.append('"' + col[row].value + '"')
            
print("Reading Completed.")
 
# Calls tableDetailsFromJson() function to fetch all tables details
tableDetailsFromJson()

# Define column headers here for output csv file
# Add all lists to data Dictionary
data["Id"] = ids
data["Name"] = names
data["ClassType"] = types
data["Schema"] = schemas
data["Business Description"] = descriptions
data["Data Owner"] = owners
data["Data Steward"] = stewards
data["AD Group"] = adgroups
data["Data Subject Area"] = subjects
data["URL"] = URLs

df = pd.DataFrame(data)

# Generate CSV file for output
df.to_csv("AssetsExport.csv", index=False)

print("*****Completed*****")
print("Generated file AssetsExport.csv")
print()
if len(inputTables) != len(tablePresent):
    tableNotPresent = list(set(inputTables) - set(tablePresent))
    if len(tableNotPresent) == 1:
        print("Below table is not present in EDC:")
    else:
        print("Below tables are not present in EDC:")
    for i in tableNotPresent:
        print(i)